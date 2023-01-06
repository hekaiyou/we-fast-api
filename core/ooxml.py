import os
import platform
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from fastapi.responses import StreamingResponse


async def get_xlsx_stream(worksheets: list):
    """ 获取临时生成的 XLSX 文件流 """
    wb = Workbook()
    ws = []
    for ws_i in range(len(worksheets)):
        # {'name': 'sheet', 'title': ['t1', 't2'], 'items': [['v1-1', 'v1-2'], ['v2-1', 'v2-2']]}
        _title = worksheets[ws_i]['name']
        for n_str in ['\\', '/', '*', '[', ']', ':', '?']:
            _title = _title.replace(n_str, '_')
        if ws_i == 0:
            ws.append(wb.active)
            ws[ws_i].title = _title
        else:
            ws.append(wb.create_sheet(title=_title))
        for t_i in range(len(worksheets[ws_i]['title'])):
            ws[ws_i].cell(
                row=1,
                column=t_i + 1,
                value=worksheets[ws_i]['title'][t_i],
            )
        for i_i in range(len(worksheets[ws_i]['items'])):
            for t_i in range(len(worksheets[ws_i]['title'])):
                ws[ws_i].cell(
                    row=i_i + 2,
                    column=t_i + 1,
                    value=worksheets[ws_i]['items'][i_i][t_i],
                )

    def temp_file():
        sys_platform = platform.platform().lower()
        if 'windows' in sys_platform:
            # delete=False 不会关闭文件后自动清理
            windows_f = NamedTemporaryFile(mode='w', delete=False)
            wb.save(windows_f.name)
            windows_f.seek(0)
            # 在 windows 下不关闭就没有权限再次打开
            windows_f.close()
            with open(windows_f.name, mode='rb') as tmp:
                yield from tmp
            # 手动清理 windows 下的临时文件
            windows_f.close()
            os.remove(windows_f.name)
        else:
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                yield from tmp

    return StreamingResponse(
        temp_file(),
        media_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
